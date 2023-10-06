# import openpyxl and tkinter modules
# from tkinter.ttk import Combobox
import tkinter as tk
from tkinter import ttk, messagebox
from openpyxl import *
from tkinter import *
import re
import tkinter as tk


wb = load_workbook('C:\\Users\\My PC\\python\\py\\lab4\\lab4\\1.xlsx')

sheet = wb.active


def excel():

    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    sheet.cell(row=1, column=1).value = "Mã số sinh viên"
    sheet.cell(row=1, column=2).value = "Họ tên"
    sheet.cell(row=1, column=3).value = "Ngày sinh"
    sheet.cell(row=1, column=4).value = "Email"
    sheet.cell(row=1, column=5).value = "Số điện thoại"
    sheet.cell(row=1, column=6).value = "Học kỳ"
    sheet.cell(row=1, column=7).value = "Năm học"

def focus0(event):
    mssv_field.focus_set()
    

def focus1(event):

    name_field.focus_set()

def focus2(event):

    ngsinh_field.focus_set()

def focus3(event):
    email_field.focus_set()

def focus4(event):
    
    sdt_field.focus_set()

def focus5(event):

    hocky_field.focus_set()

def focus6(event):

    address_field.focus_set()

def focus7(event):
    
    monhoc_field.focus_set()

def clear():

    mssv_field.delete(0, END)
    name_field.delete(0, END)
    ngsinh_field.delete(0, END)
    email_field.delete(0, END)
    sdt_field.delete(0, END)
    hocky_field.delete(0, END)
    address_field.delete(0, END)
    monhoc_field.delete(0, END)
    
# Xử lý

    # Thoát chương trình
def exit_program():
    result = messagebox.askquestion("Thoát chương trình", "Are you sure about that?")
    if result == "yes":
        root.destroy()
    # root.quit()

    # kt ngày sinh
def kiem_tra_ngay_sinh(ngsinh):
    pattern = r'^\d{2}/\d{2}/\d{4}$'
    if not re.match(pattern, ngsinh):
        messagebox.showerror("Ngày sinh không hợp lệ", "Vui lòng nhập đúng.")
        return

    ngay, thang, nam = ngsinh.split('/')
    ngay = int(ngay)
    thang = int(thang)
    nam = int(nam)

    if ngay >= 1 and ngay <= 31 and thang >= 1 and thang <= 12 and nam >= 1900 and nam <= 2099:
        return True
    return messagebox.showerror("Ngày sinh không hợp lệ", "Vui lòng nhập đúng.")

    # hocky
def kiem_tra_hoc_ky():
    hoc_ky = hocky_field.get()
    if hoc_ky not in ['1', '2', '3']:
        messagebox.showerror("Lỗi", "Học kỳ phải là 1, 2 hoặc 3. Vui lòng nhập lại.")
        return False
    return True

    # ktemail
def kiem_tra_email(email):
    pattern = r'^[\w\.-]+@[\w\.-]+\.\w+$'  # biểu thức chính quy để kiểm tra email
    if not re.match(pattern, email):
        messagebox.showerror("Email không hợp lệ","Vui lòng nhập lại")
        return False
    else:
        return True


    #ktsdt 
def ktsdt(sdt):
    if not sdt.isdigit() or len(sdt) != 10:
        messagebox.showerror("Số điện thoại không đúng", "Số điện thoại phải là số có 10 chữ số.")
        return False
    return True

    # ktmssv
def insert():
    mssv = mssv_field.get()

    if not mssv.isdigit():
        messagebox.showerror("Error", "Vui lòng nhập mã số sv!")
        return

    if len(mssv) != 7:
        messagebox.showerror("Error", "Vui lòng nhập đủ 7 số!")
        return


    if (mssv_field.get() == "" and
        name_field.get() == "" and
        ngsinh_field.get() == "" and
        email_field.get() == "" and
        sdt_field.get() == "" and
        hocky_field.get() == "" and
        address_field.get() == "" and
        monhoc_field.get() == ""):
        messagebox.showerror("Lỗi", "Vui lòng điền đầy đủ thông tin.")

        print("empty input")


    else:
        if not kiem_tra_ngay_sinh(ngsinh_field.get()):
            return
        if not kiem_tra_hoc_ky():
            return
        if not kiem_tra_email(email_field.get()):
            return
        if not ktsdt(sdt_field.get()):
            return
        current_row = sheet.max_row
        current_column = sheet.max_column

        sheet.cell(row=current_row + 1, column=1).value = mssv_field.get()
        sheet.cell(row=current_row + 1, column=2).value = name_field.get()
        sheet.cell(row=current_row + 1, column=3).value = ngsinh_field.get()
        sheet.cell(row=current_row + 1, column=4).value = email_field.get()
        sheet.cell(row=current_row + 1, column=5).value = sdt_field.get()
        sheet.cell(row=current_row + 1, column=6).value = hocky_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()
        
        
        
        wb.save('C:\\Users\\My PC\\python\\py\\lab4\\lab4\\1.xlsx')

        mssv_field.focus_set()

        clear()


if __name__ == "__main__":

    root = tk.Tk()
    
    var = IntVar()

    root.configure(background='light green')

    root.title("Đăng ký học phần")

    root.geometry("500x300")
    
    hocky_entry = tk.Entry(root)
    
    # Tạo nút kiểm tra và kết nối nó với hàm kiem_tra_hoc_ky
    kiem_tra_button = tk.Button(root, text="Kiểm tra", command=kiem_tra_hoc_ky)
    
    
    
    

    excel()

    

    heading = Label(root, text="THÔNG TIN ĐĂNG NHẬP", bg="light green")

    mssv = Label(root, text="Mã số sinh viên", bg="light green")

    name = Label(root, text="Họ tên", bg="light green")

    ngsinh = Label(root, text="Ngày sinh", bg="light green")

    email = Label(root, text="Email", bg="light green")

    sdt = Label(root, text="Số điện thoại", bg="light green")

    hocky = Label(root, text="Học kỳ", bg="light green")
    
    address = Label(root, text="Năm học", background="light green")
 
 
 
 # ############################################################nhucc
    monhoc = Label(root, text="Chọn môn học", bg="light green")
 
    

    heading.grid(row=0, column=1)
    mssv.grid(row=1, column=0)
    name.grid(row=2, column=0)
    ngsinh.grid(row=3, column=0)
    email.grid(row=4, column=0)
    sdt.grid(row=5, column=0)
    hocky.grid(row=6, column=0)
    address.grid(row=7, column=0)
    monhoc.grid(row=8, column=0)

    # entry
    mssv_field = Entry(root)
    name_field = Entry(root)
    ngsinh_field = Entry(root)
    email_field = Entry(root)
    sdt_field = Entry(root)
    hocky_field = Entry(root)	
    # combobox
    address_field = ["2022-2023", "2023-2024", "2024-2025"]
    address_field = ttk.Combobox(root, values=address_field)
    address_field.set("")
    # checklistbox
    monhoc_field = ["Lập trình","Lập trình Java"
                     "Công nghệ phầm mềm","Phát triển ứng dụng web"]
    # monhoc_field = ttk.Checkbutton(root, values=monhoc_field)
    monhoc_field = Entry(root)
    
    mssv_field.bind("<Return>", focus0)
    
    name_field.bind("<Return>", focus1)

    ngsinh_field.bind("<Return>", focus2)

    email_field.bind("<Return>", focus3)

    sdt_field.bind("<Return>", focus4)

    hocky_field.bind("<Return>", focus5)
 

    # whenever the enter key is pressed
    # then call the focus6 function
    address_field.bind("<Return>", focus6)
    monhoc_field.bind("<Return>", focus7)
 
 
    mssv_field.grid(row=1, column=1, ipadx="100")
    name_field.grid(row=2, column=1, ipadx="100")
    ngsinh_field.grid(row=3, column=1, ipadx="100")
    email_field.grid(row=4, column=1, ipadx="100")
    sdt_field.grid(row=5, column=1, ipadx="100")
    hocky_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="91")
    monhoc_field.grid(row=8,column=1, ipadx="100")



    excel()

    submit = Button(root, text="Đăng ký", fg="Black", bg="Red", command=insert,)
    submit.grid(row=9, column=1,pady=20,sticky="w",padx=70 )
    thoat = Button(root, text="Thoát", fg="Black", bg="#099900", command= exit_program)
    thoat.grid(row=9, column=1,ipadx=10)
    

    root.mainloop()
